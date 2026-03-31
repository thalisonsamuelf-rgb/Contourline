import re
import unicodedata
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

INPUT_FILE = "/Users/thalison/Documents/RDStation/analise_contatos_duplicados.md"
OUTPUT_FILE = "/Users/thalison/Documents/RDStation/analise_contatos_duplicados.xlsx"

def normalize(text):
    """Normaliza texto para comparação: minúsculo, sem acentos, sem espaços extras."""
    if not text:
        return ""
    text = text.strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return text

def normalize_phone(phone):
    """Remove tudo que não é dígito e prefixo 55."""
    if not phone:
        return ""
    digits = re.sub(r"\D", "", phone)
    # Remove prefixo 55 para normalizar
    if digits.startswith("55") and len(digits) > 11:
        digits = digits[2:]
    return digits

def parse_markdown(filepath):
    """Extrai contatos do markdown agrupados por email."""
    contacts = []
    current_email = ""
    current_section = ""  # "email" ou "nome"

    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip()

            # Detecta seção
            if line.startswith("## Duplicados por EMAIL"):
                current_section = "email"
                continue
            elif line.startswith("## Duplicados por NOME"):
                current_section = "nome"
                continue

            # Detecta email/nome do grupo
            header_match = re.match(r"^### (.+?) \(\d+x\)", line)
            if header_match:
                current_email = header_match.group(1).strip() if current_section == "email" else ""
                continue

            # Linha de tabela com dados
            if line.startswith("|") and not line.startswith("| ID") and not line.startswith("|---"):
                parts = [p.strip() for p in line.split("|")]
                parts = [p for p in parts if p != ""]
                if len(parts) >= 4:
                    contact_id = parts[0]
                    nome = parts[1]
                    telefone = parts[2]
                    criado_em = parts[3]
                    email_col = parts[4] if len(parts) > 4 else current_email

                    # Se estamos na seção de email, o email vem do header
                    if current_section == "email":
                        email_col = current_email

                    contacts.append({
                        "id": contact_id,
                        "nome": nome,
                        "telefone": telefone,
                        "email": email_col,
                        "criado_em": criado_em,
                        "secao": current_section,
                    })

    return contacts

print("Lendo arquivo...")
contacts = parse_markdown(INPUT_FILE)
print(f"Total de contatos extraídos: {len(contacts)}")

# ---- Análise de duplicação real (NOME + TELEFONE + EMAIL) ----
def chave_duplicacao(c):
    nome_n = normalize(c["nome"])
    tel_n = normalize_phone(c["telefone"])
    email_n = normalize(c["email"])
    return (nome_n, tel_n, email_n)

grupos = defaultdict(list)
for c in contacts:
    chave = chave_duplicacao(c)
    grupos[chave].append(c)

# Marca cada contato com seu status
for c in contacts:
    chave = chave_duplicacao(c)
    grupo = grupos[chave]
    if len(grupo) > 1:
        # Verifica se a chave tem campos vazios (duplicado parcial)
        nome_n, tel_n, email_n = chave
        if nome_n and tel_n and email_n:
            c["status_duplicado"] = "DUPLICADO REAL"
        elif email_n and (nome_n or tel_n):
            c["status_duplicado"] = "DUPLICADO PARCIAL"
        else:
            c["status_duplicado"] = "MESMO EMAIL"
    else:
        c["status_duplicado"] = "ÚNICO"

# ---- Criação do Excel ----
wb = Workbook()

# ---------- ABA 1: Todos os contatos ----------
ws1 = wb.active
ws1.title = "Todos os Contatos"

headers = ["ID", "Nome", "Telefone", "Email", "Criado em", "Status Duplicado", "Chave (Nome+Tel+Email)"]
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

color_map = {
    "DUPLICADO REAL": "FF4C4C",
    "DUPLICADO PARCIAL": "FFB84C",
    "MESMO EMAIL": "FFE599",
    "ÚNICO": "C6EFCE",
}

for row_idx, c in enumerate(contacts, 2):
    nome_n, tel_n, email_n = chave_duplicacao(c)
    chave_str = f"{nome_n} | {tel_n} | {email_n}"
    valores = [c["id"], c["nome"], c["telefone"], c["email"], c["criado_em"], c["status_duplicado"], chave_str]
    status = c["status_duplicado"]
    fill = PatternFill("solid", fgColor=color_map.get(status, "FFFFFF"))

    for col, val in enumerate(valores, 1):
        cell = ws1.cell(row=row_idx, column=col, value=val)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="left")

# Ajusta largura das colunas
col_widths = [28, 30, 18, 38, 14, 20, 55]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:G{len(contacts)+1}"

# ---------- ABA 2: Apenas Duplicados Reais ----------
ws2 = wb.create_sheet("Duplicados Reais")

for col, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

duplicados_reais = [c for c in contacts if c["status_duplicado"] == "DUPLICADO REAL"]
for row_idx, c in enumerate(duplicados_reais, 2):
    nome_n, tel_n, email_n = chave_duplicacao(c)
    chave_str = f"{nome_n} | {tel_n} | {email_n}"
    valores = [c["id"], c["nome"], c["telefone"], c["email"], c["criado_em"], c["status_duplicado"], chave_str]
    fill = PatternFill("solid", fgColor="FF4C4C")
    for col, val in enumerate(valores, 1):
        cell = ws2.cell(row=row_idx, column=col, value=val)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="left")

for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:G{len(duplicados_reais)+1}"

# ---------- ABA 3: Resumo por grupo ----------
ws3 = wb.create_sheet("Resumo por Grupo")

res_headers = ["Email", "Nome (normalizado)", "Telefone (normalizado)", "Qtd Registros", "Status", "IDs"]
for col, h in enumerate(res_headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

row_idx = 2
for chave, grupo in sorted(grupos.items(), key=lambda x: -len(x[1])):
    nome_n, tel_n, email_n = chave
    qtd = len(grupo)
    ids = ", ".join(c["id"] for c in grupo)

    if qtd > 1:
        if nome_n and tel_n and email_n:
            status = "DUPLICADO REAL"
            fill = PatternFill("solid", fgColor="FF4C4C")
        elif email_n and (nome_n or tel_n):
            status = "DUPLICADO PARCIAL"
            fill = PatternFill("solid", fgColor="FFB84C")
        else:
            status = "MESMO EMAIL"
            fill = PatternFill("solid", fgColor="FFE599")
    else:
        status = "ÚNICO"
        fill = PatternFill("solid", fgColor="C6EFCE")

    valores = [email_n, nome_n, tel_n, qtd, status, ids]
    for col, val in enumerate(valores, 1):
        cell = ws3.cell(row=row_idx, column=col, value=val)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="left", wrap_text=(col == 6))
    row_idx += 1

ws3.column_dimensions["A"].width = 38
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 10
ws3.column_dimensions["E"].width = 20
ws3.column_dimensions["F"].width = 80
ws3.freeze_panes = "A2"
ws3.auto_filter.ref = f"A1:F{row_idx-1}"

# ---------- ABA 4: Legenda ----------
ws4 = wb.create_sheet("Legenda")
ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 70

legenda = [
    ("COR", "SIGNIFICADO"),
    ("VERMELHO - DUPLICADO REAL", "Mesmo NOME + mesmo TELEFONE + mesmo EMAIL (normalizado). São duplicatas certas."),
    ("LARANJA - DUPLICADO PARCIAL", "Mesmo EMAIL + (mesmo nome OU mesmo telefone). Provável duplicata."),
    ("AMARELO - MESMO EMAIL", "Apenas o email é igual. Nome e telefone divergem — verificar manualmente."),
    ("VERDE - ÚNICO", "Não há outro contato com a mesma combinação nome+telefone+email."),
]

fills_legenda = ["1F4E79", "FF4C4C", "FFB84C", "FFE599", "C6EFCE"]
fonts_legenda = ["FFFFFF", "FFFFFF", "000000", "000000", "000000"]

for i, (col_a, col_b) in enumerate(legenda):
    ca = ws4.cell(row=i+1, column=1, value=col_a)
    cb = ws4.cell(row=i+1, column=2, value=col_b)
    f = PatternFill("solid", fgColor=fills_legenda[i])
    font = Font(color=fonts_legenda[i], bold=(i == 0))
    ca.fill = f; ca.font = font; ca.alignment = Alignment(horizontal="left")
    cb.fill = f; cb.font = font; cb.alignment = Alignment(horizontal="left", wrap_text=True)
    ws4.row_dimensions[i+1].height = 30

# ---- Estatísticas finais ----
total = len(contacts)
n_real = sum(1 for c in contacts if c["status_duplicado"] == "DUPLICADO REAL")
n_parcial = sum(1 for c in contacts if c["status_duplicado"] == "DUPLICADO PARCIAL")
n_email = sum(1 for c in contacts if c["status_duplicado"] == "MESMO EMAIL")
n_unico = sum(1 for c in contacts if c["status_duplicado"] == "ÚNICO")

print(f"\n=== RESUMO ===")
print(f"Total de contatos: {total}")
print(f"DUPLICADO REAL    : {n_real}")
print(f"DUPLICADO PARCIAL : {n_parcial}")
print(f"MESMO EMAIL       : {n_email}")
print(f"ÚNICO             : {n_unico}")

wb.save(OUTPUT_FILE)
print(f"\nArquivo salvo: {OUTPUT_FILE}")
